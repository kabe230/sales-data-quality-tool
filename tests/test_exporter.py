from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from sales_data_quality.exporter import formula_safe, to_csv_bytes, to_excel_bytes
from sales_data_quality.service import DataQualityService


def test_formula_injection_protection():
    assert formula_safe("  =SUM(A1:A2)") == "'  =SUM(A1:A2)"
    assert formula_safe("通常の文字") == "通常の文字"
    assert formula_safe(-100) == -100


def test_exports_have_expected_encoding_and_sheets(dataset):
    result = DataQualityService().process(dataset, execution_date=date(2026, 7, 1))
    csv_bytes = to_csv_bytes(result)
    assert csv_bytes.startswith(b"\xef\xbb\xbf")
    assert "元データ行番号" in csv_bytes.decode("utf-8-sig")

    workbook = load_workbook(BytesIO(to_excel_bytes(result)))
    assert workbook.sheetnames == [
        "概要",
        "整形済みデータ",
        "指摘一覧",
        "修正履歴",
        "ステータス別集計",
        "顧客別集計",
        "担当部署別集計",
        "登録月別集計",
        "受注予定月別集計",
    ]

    overview = workbook["概要"]
    rate_cells = {
        row[1].value: row[2]
        for row in overview.iter_rows(min_row=2)
        if row[1].value in {"正常率（%）", "有効率（%）"}
    }
    assert rate_cells["正常率（%）"].value == round(result.metrics.normal_rate, 1)
    assert rate_cells["有効率（%）"].value == round(result.metrics.valid_rate, 1)
    assert {cell.number_format for cell in rate_cells.values()} == {"0.0"}
