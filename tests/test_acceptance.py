from __future__ import annotations

from dataclasses import asdict
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from sales_data_quality.exporter import to_csv_bytes, to_excel_bytes
from sales_data_quality.loader import load_csv
from sales_data_quality.service import DataQualityService

FIXTURE = Path(__file__).parent / "fixtures" / "acceptance_projects.csv"
EXECUTION_DATE = date(2026, 7, 1)


def _process_acceptance_data():
    text = FIXTURE.read_text(encoding="utf-8")
    dataset = load_csv(f"\ufeff{text}".encode(), FIXTURE.name)
    return DataQualityService().process(dataset, execution_date=EXECUTION_DATE)


def _records(frame):
    return frame.where(frame.notna(), None).to_dict(orient="records")


def test_fixed_12_rows_match_cleaning_corrections_issues_and_metrics():
    result = _process_acceptance_data()

    assert result.cleaned_data["source_row_number"].tolist() == list(range(2, 14))
    assert result.cleaned_data.iloc[1][
        [
            "project_id",
            "customer_name",
            "contact_person",
            "email",
            "project_name",
            "amount",
            "status",
            "department",
            "registered_date",
            "expected_order_date",
        ]
    ].tolist() == [
        "P-0002",
        "株式会社みどり",
        "鈴木花子",
        "hanako@midori.co.jp",
        "クラウド移行",
        300000,
        "受注",
        "営業二部",
        date(2026, 2, 10),
        date(2026, 8, 31),
    ]
    assert result.cleaned_data.loc[5, "amount"] == "1円2"
    assert result.cleaned_data.loc[6, "registered_date"] == "2026-13-01"
    assert result.cleaned_data.loc[9, "status"] == "対応中"

    assert [
        (
            item.source_row_number,
            item.code,
            item.field_name,
            item.message,
        )
        for item in result.corrections
    ] == [
        (3, "NORMALIZED_PROJECT_ID", "project_id", "案件IDを正規化しました。"),
        (3, "TRIMMED_TEXT", "customer_name", "前後空白を除去しました。"),
        (3, "TRIMMED_TEXT", "contact_person", "前後空白を除去しました。"),
        (3, "NORMALIZED_EMAIL", "email", "メールアドレスを正規化しました。"),
        (3, "TRIMMED_TEXT", "project_name", "前後空白を除去しました。"),
        (3, "NORMALIZED_AMOUNT", "amount", "案件金額を整数へ正規化しました。"),
        (3, "NORMALIZED_STATUS", "status", "ステータスを正規化しました。"),
        (3, "TRIMMED_TEXT", "department", "前後空白を除去しました。"),
        (3, "NORMALIZED_DATE", "registered_date", "日付をISO形式へ正規化しました。"),
        (3, "NORMALIZED_DATE", "expected_order_date", "日付をISO形式へ正規化しました。"),
    ]

    assert [
        (
            item.source_row_number,
            item.severity.value,
            item.code,
            item.field_name,
            item.message,
        )
        for item in result.issues
    ] == [
        (4, "WARNING", "MISSING_DEPARTMENT", "department", "担当部署が未入力です。"),
        (5, "ERROR", "EMPTY_REQUIRED_VALUE", "customer_name", "顧客名は必須です。"),
        (6, "ERROR", "INVALID_EMAIL", "email", "メールアドレスの形式が正しくありません。"),
        (7, "ERROR", "INVALID_AMOUNT", "amount", "案件金額を整数として解釈できません。"),
        (8, "ERROR", "INVALID_DATE", "registered_date", "登録日の日付形式が正しくありません。"),
        (
            9,
            "ERROR",
            "FUTURE_REGISTERED_DATE",
            "registered_date",
            "登録日が実行基準日より未来です。",
        ),
        (
            10,
            "WARNING",
            "PAST_EXPECTED_ORDER_DATE",
            "expected_order_date",
            "受注予定日が実行基準日より過去です。",
        ),
        (
            11,
            "ERROR",
            "UNKNOWN_STATUS",
            "status",
            "ステータスが許可値または変換辞書に含まれていません。",
        ),
        (12, "ERROR", "DUPLICATE_PROJECT_ID", "project_id", "案件IDが重複しています。"),
        (13, "ERROR", "DUPLICATE_PROJECT_ID", "project_id", "案件IDが重複しています。"),
    ]

    assert result.cleaned_data[
        ["source_row_number", "check_result", "error_count", "warning_count"]
    ].to_records(index=False).tolist() == [
        (2, "NORMAL", 0, 0),
        (3, "NORMAL", 0, 0),
        (4, "WARNING", 0, 1),
        (5, "ERROR", 1, 0),
        (6, "ERROR", 1, 0),
        (7, "ERROR", 1, 0),
        (8, "ERROR", 1, 0),
        (9, "ERROR", 1, 0),
        (10, "WARNING", 0, 1),
        (11, "ERROR", 1, 0),
        (12, "ERROR", 1, 0),
        (13, "ERROR", 1, 0),
    ]
    metrics = asdict(result.metrics)
    assert metrics.pop("normal_rate") == pytest.approx(16.666666666666668)
    assert metrics.pop("valid_rate") == pytest.approx(33.333333333333336)
    assert metrics == {
        "total_rows": 12,
        "normal_rows": 2,
        "warning_rows": 2,
        "error_rows": 8,
        "aggregation_rows": 4,
        "excluded_rows": 8,
        "aggregation_amount": 2400000,
        "excluded_amount": 3450000,
        "average_amount": Decimal("600000.00"),
        "maximum_amount": 1000000,
        "minimum_amount": 300000,
        "error_issue_count": 8,
        "warning_issue_count": 2,
    }


def test_fixed_12_rows_match_all_summaries():
    summaries = _process_acceptance_data().summaries

    assert _records(summaries["status"]) == [
        {"status": "新規", "project_count": 0, "total_amount": 0, "average_amount": None},
        {"status": "商談中", "project_count": 0, "total_amount": 0, "average_amount": None},
        {"status": "見積提出", "project_count": 0, "total_amount": 0, "average_amount": None},
        {
            "status": "受注",
            "project_count": 3,
            "total_amount": 1800000,
            "average_amount": Decimal("600000.00"),
        },
        {"status": "失注", "project_count": 0, "total_amount": 0, "average_amount": None},
        {
            "status": "保留",
            "project_count": 1,
            "total_amount": 600000,
            "average_amount": Decimal("600000.00"),
        },
    ]
    assert _records(summaries["customer"]) == [
        {
            "customer_name": "株式会社さくら",
            "project_count": 1,
            "total_amount": 500000,
            "average_amount": Decimal("500000.00"),
            "maximum_amount": 500000,
            "minimum_amount": 500000,
        },
        {
            "customer_name": "株式会社みどり",
            "project_count": 1,
            "total_amount": 300000,
            "average_amount": Decimal("300000.00"),
            "maximum_amount": 300000,
            "minimum_amount": 300000,
        },
        {
            "customer_name": "株式会社未来",
            "project_count": 1,
            "total_amount": 600000,
            "average_amount": Decimal("600000.00"),
            "maximum_amount": 600000,
            "minimum_amount": 600000,
        },
        {
            "customer_name": "株式会社青空",
            "project_count": 1,
            "total_amount": 1000000,
            "average_amount": Decimal("1000000.00"),
            "maximum_amount": 1000000,
            "minimum_amount": 1000000,
        },
    ]
    assert _records(summaries["department"]) == [
        {
            "department": "営業一部",
            "project_count": 1,
            "total_amount": 1000000,
            "average_amount": Decimal("1000000.00"),
        },
        {
            "department": "営業二部",
            "project_count": 2,
            "total_amount": 900000,
            "average_amount": Decimal("450000.00"),
        },
        {
            "department": "未設定",
            "project_count": 1,
            "total_amount": 500000,
            "average_amount": Decimal("500000.00"),
        },
    ]
    assert _records(summaries["registered_month"]) == [
        {
            "registered_month": "2026-01",
            "project_count": 1,
            "total_amount": 1000000,
            "average_amount": Decimal("1000000.00"),
        },
        {
            "registered_month": "2026-02",
            "project_count": 1,
            "total_amount": 300000,
            "average_amount": Decimal("300000.00"),
        },
        {
            "registered_month": "2026-04",
            "project_count": 1,
            "total_amount": 500000,
            "average_amount": Decimal("500000.00"),
        },
        {
            "registered_month": "2026-05",
            "project_count": 1,
            "total_amount": 600000,
            "average_amount": Decimal("600000.00"),
        },
    ]
    assert _records(summaries["expected_order_month"]) == [
        {
            "expected_order_month": "2026-06",
            "project_count": 1,
            "total_amount": 600000,
        }
    ]
    assert _records(summaries["issue_breakdown"]) == [
        {
            "severity": "ERROR",
            "code": "DUPLICATE_PROJECT_ID",
            "issue_count": 2,
            "affected_row_count": 2,
        },
        {
            "severity": "ERROR",
            "code": "EMPTY_REQUIRED_VALUE",
            "issue_count": 1,
            "affected_row_count": 1,
        },
        {
            "severity": "ERROR",
            "code": "FUTURE_REGISTERED_DATE",
            "issue_count": 1,
            "affected_row_count": 1,
        },
        {
            "severity": "ERROR",
            "code": "INVALID_AMOUNT",
            "issue_count": 1,
            "affected_row_count": 1,
        },
        {
            "severity": "ERROR",
            "code": "INVALID_DATE",
            "issue_count": 1,
            "affected_row_count": 1,
        },
        {
            "severity": "ERROR",
            "code": "INVALID_EMAIL",
            "issue_count": 1,
            "affected_row_count": 1,
        },
        {
            "severity": "ERROR",
            "code": "UNKNOWN_STATUS",
            "issue_count": 1,
            "affected_row_count": 1,
        },
        {
            "severity": "WARNING",
            "code": "MISSING_DEPARTMENT",
            "issue_count": 1,
            "affected_row_count": 1,
        },
        {
            "severity": "WARNING",
            "code": "PAST_EXPECTED_ORDER_DATE",
            "issue_count": 1,
            "affected_row_count": 1,
        },
    ]


def test_fixed_12_rows_match_csv_and_excel_contracts():
    result = _process_acceptance_data()

    exported = load_csv(to_csv_bytes(result), "cleaned.csv")
    assert len(exported.dataframe) == 12
    assert exported.dataframe.columns.tolist() == [
        "案件ID",
        "顧客名",
        "担当者名",
        "メールアドレス",
        "案件名",
        "案件金額",
        "ステータス",
        "担当部署",
        "登録日",
        "受注予定日",
        "備考",
        "元データ行番号",
        "チェック結果",
        "エラー件数",
        "警告件数",
        "指摘内容",
    ]

    workbook = load_workbook(BytesIO(to_excel_bytes(result)), data_only=True)
    assert workbook.sheetnames == [
        "概要",
        "整形済みデータ",
        "指摘一覧",
        "修正履歴",
        "ステータス別集計",
        "顧客別集計",
        "担当部署別集計",
        "登録月別集計",
        "受注予定月別集計",
    ]
    assert workbook["指摘一覧"].max_row == 11
    assert workbook["修正履歴"].max_row == 11
    overview_values = [
        cell.value for row in workbook["概要"].iter_rows() for cell in row if cell.value is not None
    ]
    assert "DUPLICATE_PROJECT_ID" in overview_values
    assert "source_filename" in overview_values
