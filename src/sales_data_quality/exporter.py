from __future__ import annotations

from dataclasses import asdict
from datetime import date
from decimal import Decimal
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from sales_data_quality.config import COLUMNS, INTERNAL_DISPLAY
from sales_data_quality.models import ProcessingResult

SUMMARY_DISPLAY = {
    "status": {
        "status": "ステータス",
        "project_count": "案件件数",
        "total_amount": "合計金額",
        "average_amount": "平均金額",
    },
    "customer": {
        "customer_name": "顧客名",
        "project_count": "案件件数",
        "total_amount": "合計金額",
        "average_amount": "平均金額",
        "maximum_amount": "最大金額",
        "minimum_amount": "最小金額",
    },
    "department": {
        "department": "担当部署",
        "project_count": "案件件数",
        "total_amount": "合計金額",
        "average_amount": "平均金額",
    },
    "registered_month": {
        "registered_month": "登録月",
        "project_count": "案件件数",
        "total_amount": "合計金額",
        "average_amount": "平均金額",
    },
    "expected_order_month": {
        "expected_order_month": "受注予定月",
        "project_count": "見込案件件数",
        "total_amount": "見込金額",
    },
    "issue_breakdown": {
        "severity": "重要度",
        "code": "指摘コード",
        "issue_count": "指摘件数",
        "affected_row_count": "該当行数",
    },
}


def formula_safe(value: object) -> object:
    if not isinstance(value, str):
        return value
    probe = value.lstrip(" \u3000\t\r\n")
    return f"'{value}" if probe.startswith(("=", "+", "-", "@")) else value


def _display_data(data: pd.DataFrame) -> pd.DataFrame:
    result = data.copy(deep=True)
    result.rename(columns={**COLUMNS, **INTERNAL_DISPLAY}, inplace=True)
    result = result.map(formula_safe)
    return result


def to_csv_bytes(result: ProcessingResult) -> bytes:
    frame = _display_data(result.cleaned_data)
    text = frame.to_csv(index=False, lineterminator="\r\n", date_format="%Y-%m-%d", na_rep="")
    return text.encode("utf-8-sig")


def _issues_frame(result: ProcessingResult) -> pd.DataFrame:
    columns = [
        "元データ行番号",
        "重要度",
        "指摘コード",
        "項目",
        "修正前の値",
        "整形後の値",
        "内容",
    ]
    return pd.DataFrame(
        [
            {
                "元データ行番号": item.source_row_number,
                "重要度": item.severity.value,
                "指摘コード": item.code,
                "項目": COLUMNS[item.field_name],
                "修正前の値": formula_safe(item.original_value),
                "整形後の値": formula_safe(item.cleaned_value),
                "内容": item.message,
            }
            for item in result.issues
        ],
        columns=columns,
    )


def _corrections_frame(result: ProcessingResult) -> pd.DataFrame:
    columns = [
        "元データ行番号",
        "修正コード",
        "項目",
        "修正前の値",
        "修正後の値",
        "内容",
    ]
    return pd.DataFrame(
        [
            {
                "元データ行番号": item.source_row_number,
                "修正コード": item.code,
                "項目": COLUMNS[item.field_name],
                "修正前の値": formula_safe(item.original_value),
                "修正後の値": formula_safe(item.cleaned_value),
                "内容": item.message,
            }
            for item in result.corrections
        ],
        columns=columns,
    )


def _overview_frame(result: ProcessingResult) -> pd.DataFrame:
    metrics = asdict(result.metrics)
    labels = {
        "total_rows": "総行数",
        "normal_rows": "正常行数",
        "warning_rows": "警告行数",
        "error_rows": "エラー行数",
        "normal_rate": "正常率（%）",
        "aggregation_rows": "集計対象行数",
        "excluded_rows": "除外行数",
        "aggregation_amount": "集計対象金額",
        "excluded_amount": "除外金額",
        "average_amount": "平均金額",
        "maximum_amount": "最大金額",
        "minimum_amount": "最小金額",
        "error_issue_count": "エラー指摘数",
        "warning_issue_count": "警告指摘数",
        "valid_rate": "有効率（%）",
    }
    rows: list[dict[str, object]] = []
    for key, value in result.processing_log.items():
        rows.append({"区分": "処理基本情報", "項目": key, "値": value})
    for key, value in metrics.items():
        if key in {"normal_rate", "valid_rate"}:
            value = round(float(value), 1)
        rows.append({"区分": "基本指標", "項目": labels[key], "値": value})
    for item in result.summaries["issue_breakdown"].to_dict(orient="records"):
        rows.append(
            {
                "区分": "指摘内訳",
                "重要度": item["severity"],
                "指摘コード": item["code"],
                "指摘件数": item["issue_count"],
                "該当行数": item["affected_row_count"],
            }
        )
    return pd.DataFrame(
        rows,
        columns=["区分", "項目", "値", "重要度", "指摘コード", "指摘件数", "該当行数"],
    )


def _excel_value(value: object) -> object:
    if isinstance(value, Decimal):
        return float(value)
    return formula_safe(value)


def to_excel_bytes(result: ProcessingResult) -> bytes:
    output = BytesIO()
    sheets = {
        "概要": _overview_frame(result),
        "整形済みデータ": _display_data(result.cleaned_data),
        "指摘一覧": _issues_frame(result),
        "修正履歴": _corrections_frame(result),
        "ステータス別集計": result.summaries["status"].rename(columns=SUMMARY_DISPLAY["status"]),
        "顧客別集計": result.summaries["customer"].rename(columns=SUMMARY_DISPLAY["customer"]),
        "担当部署別集計": result.summaries["department"].rename(
            columns=SUMMARY_DISPLAY["department"]
        ),
        "登録月別集計": result.summaries["registered_month"].rename(
            columns=SUMMARY_DISPLAY["registered_month"]
        ),
        "受注予定月別集計": result.summaries["expected_order_month"].rename(
            columns=SUMMARY_DISPLAY["expected_order_month"]
        ),
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame = frame.map(_excel_value)
            frame.to_excel(writer, sheet_name=name, index=False)
            worksheet = writer.book[name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.font = Font(color="FFFFFF", bold=True)
            for column_cells in worksheet.columns:
                values = [str(cell.value or "") for cell in column_cells]
                width = min(max(len(value) for value in values) * 1.7 + 2, 45)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
            if name == "整形済みデータ":
                result_column = next(
                    (
                        cell.column
                        for cell in worksheet[1]
                        if cell.value == INTERNAL_DISPLAY["check_result"]
                    ),
                    None,
                )
                if result_column:
                    for row in range(2, worksheet.max_row + 1):
                        status = worksheet.cell(row, result_column).value
                        color = (
                            "FCE8E6"
                            if status == "ERROR"
                            else "FFF4CC"
                            if status == "WARNING"
                            else None
                        )
                        if color:
                            for cell in worksheet[row]:
                                cell.fill = PatternFill("solid", fgColor=color)
            for column_cells in worksheet.columns:
                header = column_cells[0].value
                if header and "金額" in str(header):
                    for cell in column_cells[1:]:
                        cell.number_format = "¥#,##0.00" if "平均" in str(header) else "¥#,##0"
                if header in {"登録日", "受注予定日"}:
                    for cell in column_cells[1:]:
                        if isinstance(cell.value, date):
                            cell.number_format = "yyyy-mm-dd"
            if name == "概要":
                for row in worksheet.iter_rows(min_row=2):
                    if row[1].value in {"正常率（%）", "有効率（%）"}:
                        row[2].number_format = "0.0"
    return output.getvalue()
