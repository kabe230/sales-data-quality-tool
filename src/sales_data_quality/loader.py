from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path

import pandas as pd
from charset_normalizer import from_bytes
from openpyxl import load_workbook

from sales_data_quality.config import MAX_FILE_SIZE, MAX_ROWS
from sales_data_quality.exceptions import DataQualityError
from sales_data_quality.models import LoadedDataset
from sales_data_quality.processing.cleaner import is_blank


def _check_size(content: bytes) -> None:
    if not content:
        raise DataQualityError("EMPTY_FILE", "ファイルが空です。")
    if len(content) > MAX_FILE_SIZE:
        raise DataQualityError("FILE_TOO_LARGE", "ファイルサイズが10 MiBを超えています。")


def _drop_blank_rows(dataframe: pd.DataFrame) -> tuple[pd.DataFrame, tuple[int, ...], int]:
    blank_mask = dataframe.apply(lambda row: all(is_blank(value) for value in row), axis=1)
    kept = dataframe.loc[~blank_mask].copy()
    source_rows = tuple(int(index) + 2 for index in kept.index)
    kept.reset_index(drop=True, inplace=True)
    if len(kept) > MAX_ROWS:
        raise DataQualityError("ROW_LIMIT_EXCEEDED", "データ行数が10,000行を超えています。")
    return kept, source_rows, int(blank_mask.sum())


def _validate_headers(headers: list[object]) -> None:
    normalized: list[str] = []
    for header in headers:
        if header is None:
            raise DataQualityError("EMPTY_COLUMN_NAME", "空の列名があります。")
        value = str(header).removeprefix("\ufeff").strip(" \u3000")
        if not value:
            raise DataQualityError("EMPTY_COLUMN_NAME", "空の列名があります。")
        normalized.append(value)
    if len(normalized) != len(set(normalized)):
        raise DataQualityError("DUPLICATE_COLUMN", "重複する列名があります。")


def _decode_csv(content: bytes, selected_encoding: str = "auto") -> tuple[str, str]:
    candidates = (
        [selected_encoding] if selected_encoding != "auto" else ["utf-8-sig", "utf-8", "cp932"]
    )
    for encoding in candidates:
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    if selected_encoding != "auto":
        raise DataQualityError("FILE_READ_ERROR", "指定した文字コードでCSVを読めません。")
    match = from_bytes(content).best()
    if match is None:
        raise DataQualityError("ENCODING_DETECTION_FAILED", "文字コードを判定できません。")
    return str(match), match.encoding


def load_csv(content: bytes, filename: str, encoding: str = "auto") -> LoadedDataset:
    _check_size(content)
    text, detected = _decode_csv(content, encoding)
    try:
        header = next(csv.reader(StringIO(text)))
        _validate_headers(header)
        dataframe = pd.read_csv(
            StringIO(text),
            dtype=object,
            keep_default_na=False,
            na_filter=False,
            skip_blank_lines=False,
        )
    except (StopIteration, pd.errors.EmptyDataError) as exc:
        raise DataQualityError("EMPTY_FILE", "ファイルが空です。") from exc
    except DataQualityError:
        raise
    except Exception as exc:
        raise DataQualityError("FILE_READ_ERROR", "CSVを読み込めません。") from exc
    dataframe, source_rows, blank_count = _drop_blank_rows(dataframe)
    return LoadedDataset(
        dataframe,
        Path(filename).name,
        "csv",
        detected,
        None,
        source_rows,
        tuple(str(column) for column in dataframe.columns),
        len(content),
        blank_count,
    )


def list_excel_sheets(content: bytes) -> list[str]:
    _check_size(content)
    try:
        return pd.ExcelFile(BytesIO(content), engine="openpyxl").sheet_names
    except Exception as exc:
        raise DataQualityError("FILE_READ_ERROR", "Excelを読み込めません。") from exc


def load_excel(content: bytes, filename: str, sheet_name: str) -> LoadedDataset:
    _check_size(content)
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise DataQualityError("SHEET_NOT_FOUND", "指定したシートが見つかりません。")
        header = [cell.value for cell in next(workbook[sheet_name].iter_rows(max_row=1))]
        _validate_headers(header)
        workbook.close()
        dataframe = pd.read_excel(
            BytesIO(content), sheet_name=sheet_name, dtype=object, engine="openpyxl"
        )
    except DataQualityError:
        raise
    except ValueError as exc:
        raise DataQualityError("SHEET_NOT_FOUND", "指定したシートが見つかりません。") from exc
    except Exception as exc:
        raise DataQualityError("FILE_READ_ERROR", "Excelを読み込めません。") from exc
    dataframe, source_rows, blank_count = _drop_blank_rows(dataframe)
    return LoadedDataset(
        dataframe,
        Path(filename).name,
        "xlsx",
        None,
        sheet_name,
        source_rows,
        tuple(str(column) for column in dataframe.columns),
        len(content),
        blank_count,
    )


def load_file(
    content: bytes,
    filename: str,
    *,
    encoding: str = "auto",
    sheet_name: str | None = None,
) -> LoadedDataset:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return load_csv(content, filename, encoding)
    if suffix == ".xlsx":
        sheets = list_excel_sheets(content)
        return load_excel(content, filename, sheet_name or sheets[0])
    raise DataQualityError("UNSUPPORTED_FILE_TYPE", "CSVまたは.xlsxを選択してください。")
